from openpyxl import load_workbook
import os
import sys

basePath = "timecard_input/"
outputPath = "timecard_output/"

PAYINFO_NAME_COLUMN = 2
PAYINFO_START_ROW = 2
PAYINFO_START_COLUMN = 3
PAYINFO_ITEM_NUM = 7
TIMERECORDS_NAME_COLUMN = 1
TIMERECORDS_TIME_COLUMN = 6
TIMERECORDS_PRJ_COLUMN = 26
TIMERECORDS_CONTRACT_COLUMN = 19


def init():
    # load all source excel from specified dirs
    sources = os.listdir(basePath)
    print(sources)
    # process all excels under this folder
    for i in range(0, len(sources)):
        fileName = sources[i]
        if not fileName.endswith('.xlsx') or fileName.startswith('~'):
            continue
        print("start process:{}".format(fileName))
        wb = load_workbook(basePath + fileName)
        processSource(wb, fileName)


def loadPayInfo(sheet):
    print("开始预处理薪资表 {}\r".format(sheet.title))
    result = {}
    for i in range(PAYINFO_START_ROW, sheet.max_row + 1):
        print("\r预处理薪资：{}".format(i), end='')
        nameCell = sheet.cell(row=i, column=PAYINFO_NAME_COLUMN)
        if isEmptyCell(nameCell):
            print("\r发现空行，结束读表")
            return result
        name = nameCell.value
        result[name] = []
        for j in range(PAYINFO_START_COLUMN, PAYINFO_ITEM_NUM + 3):
            try:
                payItem = float(sheet.cell(row=i, column=j).value)
            except Exception as e:
                print("\r薪资表包含无效数字：{} {}".format(name, sheet.cell(row=i, column=j).value))
                sys.exit(0)
            result[name].append(payItem)
    # print(result)
    print("\r预处理薪资表结束\r")
    return result


def isEmptyCell(cell):
    return cell.value is None or len(str(cell.value).strip()) == 0


def updateResultDict(prjOrContractResult, employeeResult, prjType, name, code, hours):
    code = str(code)
    if code not in prjOrContractResult:
        prjOrContractResult[code] = {
            "hours": hours,
            "cost": [0.0] * PAYINFO_ITEM_NUM
        }
    else:
        prjOrContractResult[code]["hours"] += hours

    # update employee store
    prjTypeKey = "prjDict"
    if prjType == 1:
        prjTypeKey = "contractDict"

    if code not in employeeResult[name][prjTypeKey]:
        employeeResult[name][prjTypeKey][code] = hours
    else:
        employeeResult[name][prjTypeKey][code] += hours


def processSource(wb, fileName):
    # result dict
    projectResultDict = {}
    contractResultDict = {}
    # some count store, dict key is employee name, store total working hours
    employeeDict = {}
    timeRecordsSheet = wb[wb.sheetnames[0]]
    contractPayInfosSheet = wb[wb.sheetnames[1]]
    projectPayInfosSheet = wb[wb.sheetnames[2]]
    # load contract pay info
    contractPayInfo = loadPayInfo(contractPayInfosSheet)
    # load project pay info
    projectPayInfo = loadPayInfo(projectPayInfosSheet)
    # go through timeRecordsSheet row by row
    for i in range(2, timeRecordsSheet.max_row + 1):
        nameCell = timeRecordsSheet.cell(row=i, column=TIMERECORDS_NAME_COLUMN)
        name = nameCell.value;
        print("\r预处理：{} {}".format(i, name), end='')
        if isEmptyCell(nameCell):
            print("\r该表预处理完成")
            break
        # check project or contract
        prjCell = timeRecordsSheet.cell(row=i, column=TIMERECORDS_PRJ_COLUMN)
        contractCell = timeRecordsSheet.cell(row=i, column=TIMERECORDS_CONTRACT_COLUMN)
        try:
            hourCell = timeRecordsSheet.cell(row=i, column=TIMERECORDS_TIME_COLUMN)
            hours = float(hourCell.value)
        except Exception as e:
            print("\r忽略无效工时记录：{}".format(i, name, hourCell.value))
            continue

        if name not in employeeDict:
            employeeDict[name] = {
                "totalHours": hours,
                "prjDict": {},
                "contractDict": {}
            }
        else:
            employeeDict[name]["totalHours"] += hours

        # consider prj first then contract, permit them exist both
        if not isEmptyCell(prjCell):
            code = prjCell.value
            updateResultDict(projectResultDict, employeeDict, 0, name, code, hours)
        elif not isEmptyCell(contractCell):
            code = contractCell.value
            updateResultDict(contractResultDict, employeeDict, 1, name, code, hours)

    # store error record
    missPrjName = set()
    missContractName = set()
    # now start calculate
    # go through employee dict
    for name in employeeDict:
        prjDict = employeeDict[name]['prjDict']
        contractDict = employeeDict[name]['contractDict']
        employeeMonthHours = employeeDict[name]['totalHours']

        def updateResult(codeInter, dictOrg, dictUpdate, payInfo, miss):
            prjHours = dictOrg[codeInter]
            rate = prjHours / employeeMonthHours

            if name not in payInfo:
                miss.add(name)
                return False

            compute = [p * rate for p in payInfo[name]]
            dictUpdate[codeInter]['cost'] = [round(a + b, 2) for a, b in zip(dictUpdate[codeInter]['cost'], compute)]
            return True

        for prjCode in prjDict:
            updateResult(prjCode, prjDict, projectResultDict, projectPayInfo, missPrjName)

        for contractCode in contractDict:
            updateResult(contractCode, contractDict, contractResultDict, contractPayInfo, missContractName)

    print("\n--------------------------------")
    if len(missPrjName) > 0:
        print("{} 缺少以下研发人员薪资信息：{}".format(fileName, missPrjName))
    if len(missContractName) > 0:
        print("{} 缺少以下实施人员薪资信息：{}".format(fileName, missContractName))

    # write out result excel
    resultPrjSheetName = "项目汇总"
    resultContractSheetName = "实施汇总"

    if resultPrjSheetName in wb.sheetnames:
        del wb[resultPrjSheetName]
    if resultContractSheetName in wb.sheetnames:
        del wb[resultContractSheetName]

    updateTarget(wb, resultPrjSheetName, projectResultDict)
    updateTarget(wb, resultContractSheetName, contractResultDict)
    wb.save(outputPath + fileName)
    print("\n输出结果至：{}".format(outputPath + fileName))


def updateTarget(wb, title, dict):
    targetSheet = wb.create_sheet(title=title)
    targetSheet.append(["代号", "总工时", "个人工资", "单位养老", "单位失业", "单位工伤", "单位生育", "单位医疗", "单位公积金"])
    rows = sorted(dict.items(), key=lambda item: item[0])
    sumRow = ['累加', 0.0] + [0.0] * PAYINFO_ITEM_NUM

    for row in rows:
        print("\r生成结果：{}".format(row[0]), end='')
        cell_row = [row[0], row[1]['hours']]
        cell_row += row[1]['cost']
        targetSheet.append(cell_row)
        sumRow[1] += row[1]['hours']
        for i in range(0, PAYINFO_ITEM_NUM):
            sumRow[i + 2] += row[1]['cost'][i]

    targetSheet.append(sumRow)


init()
