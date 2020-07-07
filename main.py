from openpyxl import load_workbook
from dateutil import relativedelta
from datetime import datetime

wb = load_workbook("input.xlsx")

print(wb.sheetnames, '\n')

targetSheet = wb[wb.sheetnames[0]]
projectSheet = wb[wb.sheetnames[1]]
salarySheet = wb[wb.sheetnames[2]]

# project hash map
project_fee_map = {}

# get count period from salary sheet
def getSalaryStartEnd():
    start = str(salarySheet.cell(row=1, column=2).value)
    end = str(salarySheet.cell(row=1, column=salarySheet.max_column).value)

    start = datetime.strptime(start, '%Y%m')
    end = datetime.strptime(end, '%Y%m')
    return (start, end)

def locateEmployee(name):
    for row in range(2, salarySheet.max_row + 1):
        if salarySheet.cell(row=row, column=1).value == name:
            return row
    return None

def locateProjectStartIdx(startDate, endDate):
    (validStart, validEnd) = getSalaryStartEnd()

    # month based shift
    salarySheetStartColum = 2
    salarySheetEndColum = salarySheet.max_column

    if startDate > validStart:
        startDiff = relativedelta.relativedelta(startDate, validStart)
        salarySheetStartColum += startDiff.months + (12 * startDiff.years)

    if endDate < validEnd:
        endDiff = relativedelta.relativedelta(validEnd, endDate)
        salarySheetEndColum -= endDiff.months + (12 * endDiff.years) + 1

    #validateion check
    if 0 < salarySheetStartColum <= salarySheet.max_column and salarySheetStartColum <= salarySheetEndColum <= salarySheet.max_column:
        return salarySheetStartColum, salarySheetEndColum
    else:
        return 0, 0


# go through excel to retrive
for i in range(5, projectSheet.max_row + 1):
    nameCell = projectSheet.cell(row=i, column=2)
    print(nameCell.value)

    if nameCell.value is None:
       continue

    salaryShare = [0] * (salarySheet.max_column + 1)
    #first column iteration to count month salary share
    for fi in range(5, projectSheet.max_column + 1):
        projectActionStatus = projectSheet.cell(row=i, column=fi).value
        if isinstance(projectActionStatus, str):
            projectActionStatus = projectActionStatus.strip();

        if not projectActionStatus is None and projectActionStatus != "":
            # start count
            # get salary info of project start to end
            startDate = projectSheet.cell(row=3, column=fi).value
            endDate = projectSheet.cell(row=4, column=fi).value

            (cS, cE) = locateProjectStartIdx(startDate, endDate)
            if cS != 0:
                monthes = [];
                if projectActionStatus != 0:
                    monthes = set(map(lambda x: int(x) + 1, projectActionStatus.split(" ")))


                for csIdx in range(cS, cE + 1):
                    # check only specify month index case
                    if projectActionStatus == 0 or csIdx in monthes:
                        salaryShare[csIdx] += 1



    # get project id
    for j in range(5, projectSheet.max_column + 1):
        projectNo = projectSheet.cell(row=2, column=j).value
        projectActionStatus = projectSheet.cell(row=i, column=j).value
        if isinstance(projectActionStatus, str):
            projectActionStatus = projectActionStatus.strip();

        emName = projectSheet.cell(row=i, column=2).value
        if not projectActionStatus is None and projectActionStatus != "":
            # start count
            # get salary info of project start to end
            startDate = projectSheet.cell(row=3, column=j).value
            endDate = projectSheet.cell(row=4, column=j).value

            (cS, cE) = locateProjectStartIdx(startDate, endDate)

            if cS != 0:
                salaryCount = 0
                print("emName:", emName)
                sRowIndx = locateEmployee(emName)

                monthes = [];
                if projectActionStatus != 0:
                    monthes = set(map(lambda x: int(x) + 1, projectActionStatus.split(" ")))

                for k in range(cS, cE + 1):
                    monthSalary = salarySheet.cell(row=sRowIndx, column=k).value
                    if monthSalary is None:
                        monthSalary = 0
                    actualShare = monthSalary / salaryShare[k]

                    if projectActionStatus == 0 or k in monthes:
                        if not projectNo in project_fee_map:
                            project_fee_map[projectNo] = [0] * 36 # 3years

                        project_fee_map[projectNo][k] += actualShare

            #print("account:", project_fee_map[projectNo])

#print(project_fee_map)

# apply to sheet 1
for i in range(5, targetSheet.max_row + 1):
    for j in range(2, targetSheet.max_column + 1):
        projectNo = targetSheet.cell(row=2, column=j).value
        data = project_fee_map[projectNo]
        targetSheet.cell(row=i, column=j).value = data[i - 5 + 2]


wb.save("out.xlsx")
print("process done")